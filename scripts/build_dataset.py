#!/usr/bin/env python3
"""
日本食品標準成分表（八訂）増補2023年 のExcelファイル群を、
Cloudflare Pagesで配信する静的JSON API (index.json / data/{食品番号}.json /
metadata.json) にビルドするスクリプト。

data/ 配下のExcelは複数行の結合ヘッダーを持つ「神エクセル」なので、
各ファイル共通の「成分識別子」行・「単位」行を機械的に検出し、
食品番号（5桁）をキーにデータ行を抽出する。

再実行可能（public/ を都度作り直す）。
"""
import glob
import json
import os
import re

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, "data")
OUT_DIR = os.path.join(ROOT, "public")
OUT_DATA_DIR = os.path.join(OUT_DIR, "data")

CODE_RE = re.compile(r"^\d{5}$")

# (ファイル番号サフィックス, カテゴリパス, テーブル名, 出典タイトル)
TABLES = [
    ("02", ("general",), "本表",
     "日本食品標準成分表（八訂）増補2023年 本表"),
    ("04", ("amino_acids", "per_100g"), "アミノ酸成分表 第1表",
     "アミノ酸成分表編 第1表（可食部100g当たり）"),
    ("05", ("amino_acids", "per_g_nitrogen"), "アミノ酸成分表 第2表",
     "アミノ酸成分表編 第2表（基準窒素1g当たり）"),
    ("06", ("amino_acids", "per_g_protein"), "アミノ酸成分表 第3表",
     "アミノ酸成分表編 第3表（アミノ酸組成によるたんぱく質1g当たり）"),
    ("07", ("amino_acids", "per_g_protein_nitrogen"), "アミノ酸成分表 第4表",
     "アミノ酸成分表編 第4表（基準窒素によるたんぱく質1g当たり）"),
    ("09", ("fatty_acids", "per_100g"), "脂肪酸成分表 第1表",
     "脂肪酸成分表編 第1表（可食部100g当たり・本表）"),
    ("10", ("fatty_acids", "per_100g_fatty_acid"), "脂肪酸成分表 第2表",
     "脂肪酸成分表編 第2表（脂肪酸100g当たり・脂肪酸組成表）"),
    ("11", ("fatty_acids", "per_g_lipid"), "脂肪酸成分表 第3表",
     "脂肪酸成分表編 第3表（脂質1g当たり）"),
    ("13", ("carbohydrates", "available_and_polyols"), "炭水化物成分表 本表",
     "炭水化物成分表編 本表（利用可能炭水化物・糖アルコール）"),
    ("14", ("carbohydrates", "fiber"), "炭水化物成分表 別表1",
     "炭水化物成分表編 別表1（食物繊維）"),
    ("15", ("organic_acids",), "炭水化物成分表 別表2",
     "炭水化物成分表編 別表2（有機酸）"),
]

FOOD_GROUPS = {
    "01": "穀類", "02": "いも及びでん粉類", "03": "砂糖及び甘味類", "04": "豆類",
    "05": "種実類", "06": "野菜類", "07": "果実類", "08": "きのこ類", "09": "藻類",
    "10": "魚介類", "11": "肉類", "12": "卵類", "13": "乳類", "14": "油脂類",
    "15": "菓子類", "16": "し好飲料類", "17": "調味料及び香辛料類",
    "18": "調理済み流通食品類",
}


def norm(s):
    """全角/半角空白・改行を除去した比較用文字列"""
    if s is None:
        return None
    return re.sub(r"[\s　]+", "", str(s))


def clean_unit(s):
    """単位行に含まれる版組み用の点線・丸括弧・空白を除去する"""
    if s is None:
        return None
    s = re.sub(r"[（）()]", "", str(s))
    s = re.sub(r"[…\.．・]+", "", s)
    s = re.sub(r"[\s　]+", "", s)
    return s


def find_file(suffix):
    matches = glob.glob(os.path.join(DATA_DIR, f"*_{suffix}.xlsx"))
    if not matches:
        raise FileNotFoundError(f"table suffix {suffix} not found in {DATA_DIR}")
    return matches[0]


def merged_value(ws, merged_ranges, row, col):
    """結合セルの場合は左上セルの値をフォワードフィルして返す"""
    val = ws.cell(row=row, column=col).value
    if val is not None:
        return val
    for rng in merged_ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return None


def find_header_rows(ws, max_col):
    """「成分識別子」行と「単位」行、および「備考」列を検出する"""
    id_row = unit_row = None
    remarks_col = None
    for r in range(1, 16):
        v = norm(ws.cell(row=r, column=4).value)
        if v == "成分識別子":
            id_row = r
        if v == "単位":
            unit_row = r
    if id_row is None or unit_row is None:
        raise ValueError("id_row/unit_row not found")
    header_top = min(id_row, unit_row)
    for r in range(1, header_top + 1):
        for c in range(5, max_col + 1):
            if norm(ws.cell(row=r, column=c).value) == "備考":
                remarks_col = c
                break
        if remarks_col:
            break
    return id_row, unit_row, remarks_col


def build_labels(ws, merged_ranges, header_top, header_bottom, max_col):
    """列ごとに、ヘッダー行を結合セルフォワードフィルしながら積み上げた和名ラベルを作る"""
    # 各行を forward-fill した値のリストを作る
    rows_values = []
    for r in range(header_top, header_bottom + 1):
        rows_values.append([merged_value(ws, merged_ranges, r, c) for c in range(1, max_col + 1)])

    # 各行について、半数以上の列で同一値になっている「バナー値」（例：「可食部100g当たり」）を検出し、
    # ラベル生成時にはその値を除外する（1列だけ違う値、例：「備考」列見出しは通常どおり残す）
    row_banner_norm = []
    for row_vals in rows_values:
        seg = row_vals[4:]
        non_none = [x for x in seg if x is not None]
        norm_counts = {}
        for x in non_none:
            nx = re.sub(r"[\s　]+", "", str(x))
            norm_counts[nx] = norm_counts.get(nx, 0) + 1
        banner = None
        if non_none:
            top_val, top_count = max(norm_counts.items(), key=lambda kv: kv[1])
            if top_count >= max(2, len(non_none) * 0.5):
                banner = top_val
        row_banner_norm.append(banner)

    labels = {}
    for c in range(5, max_col + 1):
        idx = c - 1  # 0-based column index (matches identifier map key)
        parts = []
        for row_vals, banner in zip(rows_values, row_banner_norm):
            v = row_vals[c - 1]
            if v is None:
                continue
            text = re.sub(r"[\s　]+", "", str(v))
            if banner is not None and text == banner:
                continue
            if text and (not parts or parts[-1] != text):
                parts.append(text)
        labels[idx] = "－".join(parts)
    return labels


def normalize_value(raw):
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return {"raw": raw, "value": raw, "estimated": False, "trace": False, "unmeasured": False}
    s = str(raw).strip()
    if s == "":
        return {"raw": "", "value": None, "estimated": False, "trace": False, "unmeasured": False}
    estimated = False
    inner = s
    m = re.match(r"^\((.*)\)$", s)
    if m:
        estimated = True
        inner = m.group(1).strip()
    trace = inner == "Tr"
    unmeasured = inner == "-"
    value = None
    if not trace and not unmeasured and inner != "":
        try:
            value = float(inner)
            if value == int(value):
                value = int(value)
        except ValueError:
            value = None
    return {"raw": s, "value": value, "estimated": estimated, "trace": trace, "unmeasured": unmeasured}


def set_nested(rec, path, value):
    target = rec
    for p in path[:-1]:
        target = target.setdefault(p, {})
    target[path[-1]] = value


def main():
    os.makedirs(OUT_DATA_DIR, exist_ok=True)

    records = {}
    identifiers_meta = {}  # identifier -> {label, unit, table, category, source}

    for suffix, path, table_name, source_title in TABLES:
        file_path = find_file(suffix)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["表全体"]
        max_col = ws.max_column
        max_row = ws.max_row
        merged_ranges = list(ws.merged_cells.ranges)

        id_row, unit_row, remarks_col = find_header_rows(ws, max_col)
        header_top = 2
        header_bottom = min(id_row, unit_row) - 1
        labels = build_labels(ws, merged_ranges, header_top, header_bottom, max_col)

        col_id_map = {}
        for c in range(5, max_col + 1):
            ident = ws.cell(row=id_row, column=c).value
            if ident is None:
                continue
            ident = str(ident).strip()
            if ident:
                col_id_map[c - 1] = ident

        unit_map = {}
        for c in range(5, max_col + 1):
            u = merged_value(ws, merged_ranges, unit_row, c)
            if u is not None:
                unit_map[c - 1] = clean_unit(u)

        remarks_idx = (remarks_col - 1) if remarks_col else None
        category_str = ".".join(path)

        for ident, idx in col_id_map.items():
            pass  # placeholder (kept for readability of loop separation below)

        for idx, ident in col_id_map.items():
            if ident not in identifiers_meta:
                identifiers_meta[ident] = {
                    "label": labels.get(idx, ""),
                    "unit": unit_map.get(idx),
                    "table": table_name,
                    "category": category_str,
                    "source": source_title,
                }

        data_start = max(id_row, unit_row) + 1
        for r in range(data_start, max_row + 1):
            code = ws.cell(row=r, column=2).value
            if code is None or not CODE_RE.match(str(code).strip()):
                continue
            code = str(code).strip()
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]

            rec = records.setdefault(code, {"code": code})
            group = str(row_vals[0]).strip() if row_vals[0] is not None else code[:2]
            rec.setdefault("group", group)
            rec.setdefault("group_name", FOOD_GROUPS.get(group, ""))
            index_no = row_vals[2]
            if index_no is not None:
                rec.setdefault("index_no", str(index_no).strip())
            name = row_vals[3]
            if name and not rec.get("name"):
                rec["name"] = str(name).strip()

            values = {}
            footnotes = []
            for c in range(5, max_col + 1):
                idx = c - 1
                raw = row_vals[c - 1]
                if idx in col_id_map:
                    nv = normalize_value(raw)
                    if nv is not None:
                        values[col_id_map[idx]] = nv
                elif idx != remarks_idx:
                    if raw not in (None, ""):
                        footnotes.append({"column": idx, "raw": raw})
            if footnotes:
                values["_footnote_marks"] = footnotes

            set_nested(rec, path, values)

            if remarks_idx is not None:
                rv = row_vals[remarks_idx]
                if rv not in (None, ""):
                    rec.setdefault("remarks", {})[category_str] = str(rv)

        wb.close()
        print(f"[{suffix}] {table_name}: {len(col_id_map)} identifiers, data rows scanned up to row {max_row}")

    # --- 出力 ---
    all_codes = sorted(records.keys())
    print(f"total unique food codes: {len(all_codes)}")

    index_list = []
    for code in all_codes:
        rec = records[code]
        index_list.append({
            "code": code,
            "index_no": rec.get("index_no"),
            "group": rec.get("group"),
            "group_name": rec.get("group_name"),
            "name": rec.get("name"),
        })
        with open(os.path.join(OUT_DATA_DIR, f"{code}.json"), "w", encoding="utf-8") as f:
            json.dump(rec, f, ensure_ascii=False, separators=(",", ":"))

    with open(os.path.join(OUT_DIR, "index.json"), "w", encoding="utf-8") as f:
        json.dump(index_list, f, ensure_ascii=False, separators=(",", ":"))

    metadata = {
        "title": "食品成分表示API",
        "source": "日本食品標準成分表（八訂）増補2023年（文部科学省）",
        "edition_note": "data/ 配下のExcel原本のファイル名日付をそのまま採用",
        "food_count": len(all_codes),
        "food_groups": FOOD_GROUPS,
        "tables": [
            {"category": ".".join(path), "table_name": name, "source": title}
            for _, path, name, title in TABLES
        ],
        "identifiers": identifiers_meta,
        "value_annotation_legend": {
            "raw": "Excel原本のセルの値をそのまま文字列/数値で保持したもの",
            "value": "raw を数値化できた場合の数値。Tr・-・未測定・空欄などは null",
            "estimated": "true の場合、raw が「(数値)」の形式＝他の成分値等からの推計値であることを示す",
            "trace": "true の場合、raw が「Tr」＝微量（最小記載量に達していないが含まれる）であることを示す",
            "unmeasured": "true の場合、raw が「-」＝未測定であることを示す",
            "_footnote_marks": "識別子の割り当てられていない結合セル（脚注記号等）にゴミでない値があった場合にのみ、column（0始まり列インデックス）とraw値を保持する",
        },
        "primary_key": "code（食品番号、5桁）",
        "endpoints": {
            "index": "/index.json",
            "item": "/data/{食品番号}.json",
            "metadata": "/metadata.json",
            "usage_guide": "/llms.txt",
        },
    }
    with open(os.path.join(OUT_DIR, "metadata.json"), "w", encoding="utf-8") as f:
        json.dump(metadata, f, ensure_ascii=False, indent=1)

    write_llms_txt(len(all_codes))

    print(f"wrote {len(index_list)} data files + index.json + metadata.json + llms.txt to {OUT_DIR}")


def write_llms_txt(food_count):
    content = f"""# 食品成分表示API

> 文部科学省「日本食品標準成分表（八訂）増補2023年」を元にした、食品の成分値を返す静的JSON API。食品番号をキーに、一般成分・アミノ酸・脂肪酸・炭水化物（利用可能炭水化物／食物繊維／有機酸）を横断的に取得できる。

全{food_count}食品のデータを収録。Cloudflare Pages上の静的ファイルとして配信しており、認証は不要。

## 使い方（推奨手順）

1. `/index.json` を取得し、`name`（食品名）で部分一致検索するなどして目的の食品の `code`（食品番号、5桁）を特定する。
2. `/data/{{code}}.json` を取得し、成分データを得る（例: `/data/01001.json`）。
3. 各成分値の識別子（`WATER`, `ENERC` など）の日本語ラベル・単位・出典テーブルは `/metadata.json` の `identifiers` を参照する。

## 値の読み方

各成分値は次の形のオブジェクトになっている:

```json
{{"raw": "(11.3)", "value": 11.3, "estimated": true, "trace": false, "unmeasured": false}}
```

- `raw`: Excel原本のセルの値をそのまま保持したもの
- `value`: 数値化できた場合の数値。`Tr`・`-`・空欄などは `null`
- `estimated`: `true` なら推計値（他の成分値等からの計算値。原本で `(数値)` と表記）
- `trace`: `true` なら微量（`Tr`。最小記載量未満だが含まれる）
- `unmeasured`: `true` なら未測定（`-`）

食品によっては、識別子の割り当てられていない結合セルに脚注記号（`*`等）が入っている場合があり、その場合は各カテゴリの `_footnote_marks` 配列にそのまま保持している。

## エンドポイント

- [/index.json](/index.json): 全{food_count}食品の `{{code, index_no, group, group_name, name}}` 一覧
- [/data/{{食品番号}}.json](/data/01001.json): 1食品の成分データ（例として01001＝アマランサス 玄穀）
- [/metadata.json](/metadata.json): 識別子辞書・食品群コード表・値注記の凡例・出典テーブル一覧
- /llms.txt: このファイル

## 出典・注意事項

出典は文部科学省「日本食品標準成分表（八訂）増補2023年」。一般成分表（本表）は可食部100 g当たりの値。アミノ酸成分表・脂肪酸成分表は複数の基準（可食部100 g当たり／たんぱく質・脂質1 g当たり／基準窒素1 g当たり等）があり、`/metadata.json` の `identifiers[].category` と `unit` でどの基準かを確認できる。
"""
    with open(os.path.join(OUT_DIR, "llms.txt"), "w", encoding="utf-8") as f:
        f.write(content)


if __name__ == "__main__":
    main()
